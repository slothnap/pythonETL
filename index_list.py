# python index_list.py

# import DbConfig as dbConf
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook



if __name__ == '__main__':
    # MDM DB 연결
    # AppDB = dbConf.DbConfig("app_data_export")
    # AppDB.opendb()
    # print("접속완료")
    #
    # # cast(aes_decrypt(unhex(db.user_pwd), 'jasonmdm' ) as char) as password\n"
    # sql = ("""
    #        select TABLE_SCHEMA, TABLE_NAME as table_name, INDEX_NAME -- , SEQ_IN_INDEX, COLUMN_NAME
    #             , GROUP_CONCAT(DISTINCT COLUMN_NAME ORDER BY SEQ_IN_INDEX asc SEPARATOR ', ') as index_columns
    #          from information_schema.statistics
    #         where TABLE_SCHEMA = 'jasonapp018'
    #           and table_name in ('ord_order','ord_item')
    #         group by TABLE_SCHEMA, TABLE_NAME, INDEX_NAME
    #         order by TABLE_NAME asc;
    #        """
    #        )

    # DataFrame으로 만들기
    #df = pd.DataFrame(AppDB.select(sql))
    #print(df)

    # 테이블 별로 시트 저장
    # with pd.ExcelWriter('C:/Users/ChoiYouJin/Desktop/ce1.xlsx') as writer:
    #     for name in df.table_name:
    #         df[df.table_name == f'{name}'].to_excel(writer, sheet_name=f'{name}')

    # load_wb = load_workbook("C:/Users/ChoiYouJin/Desktop/업무/8. 요일별/인덱스정의서_템플릿.xlsx")
    # load_ws = load_wb['ord_order']

    load_wb = load_workbook("C:/Users/ChoiYouJin/Desktop/업무/8. 요일별/인덱스정의서_템플릿.xlsx", data_only=True)
    load_ws = load_wb['temp']

    write_wb = Workbook()
    write_ws = write_wb.active

    # 실제값 넣기
    write_ws['A1'] = load_ws['A1'].value
    write_ws['A2'] = load_ws['A2'].value
    write_ws['E2'] = load_ws['E2'].value
    write_ws['I2'] = load_ws['I2'].value
    write_ws['A2'] = load_ws['A2'].value


    write_wb.save("C:/Users/ChoiYouJin/Desktop/업무/8. 요일별/인덱스.xlsx")


    # AppDB.closedb()

